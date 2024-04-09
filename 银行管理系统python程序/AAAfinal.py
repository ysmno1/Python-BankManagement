import threading
from tkinter import *
import tkinter
from openpyxl import  load_workbook
from tkinter.ttk import Treeview
import tkinter.messagebox as msgbox
from random import *
import pyperclip as pc
from PIL import Image,ImageTk
#窗口居中
def cent(root,w,h):
    screenw=root.winfo_screenwidth()
    screenh=root.winfo_screenheight()
    x=(screenw-w)/2
    y=(screenh-h)/2
    root.geometry('%dx%d+%d+%d'%(w,h,x,y))

#Get image
def get_image(root,file,height,width):
    global image
    img=Image.open(file).resize((width,height))
    image=ImageTk.PhotoImage(img,master=root)
    return image

#获取GIF
def get_gif(numIdx,st,file,master=None):
    frames = [PhotoImage(master=master,file=file, format='gif -index %i' %(i)) for i in range(st,numIdx+st)]
    return frames

#打开文件
wb=load_workbook("Mytest.xlsx")
ws=wb['mytest']
ws1=wb['data']
i=0
#汇率
def exrate(event):
    global RMB_entry
    root5=Tk()
    root5.resizable(0,0)
    root5.title('欢迎登录')
    root5.geometry('600x230')
    Label(root5, text="汇率更新于7月4日",font=("宋体",20)).place(x=300,y=180)
    RMB = Label(root5, text="人民币面额",font=("宋体",20))
    RMB.place(x=30, y=30)
    RMB_entry=Entry(root5,font=("宋体",20))
    RMB_entry.place(x=200, y=30)
    Button_dollar = Button(root5, text="美元",font=('宋体',20),command=dollar)
    Button_dollar.place(x=50, y=120)
    Button_pound = Button(root5, text="英镑",font=('宋体',20),command=pound)
    Button_pound.place(x=250, y=120)
    Button_yen = Button(root5, text="日元", font=('宋体',20),command=yen)
    Button_yen.place(x=450, y=120)   
def dollar():
    rmb=RMB_entry.get()
    while True:
        try:
            rmb=float(rmb)
            rmb=round(rmb,2)
        except:
            msgbox.showinfo('提示','非法输入')
            break   
        if rmb>0:
            root6=Tk()
            root6.resizable(0,0)
            root6.title("美元")
            root6.geometry("300x150")
            ammount=float(rmb)/6.70
            Label_dollar = Label(root6, text="可兑换%.2f美元"%(ammount),font=('宋体',20))
            Label_dollar.pack()
            Button(root6, text="确定",command=root6.destroy,font=('宋体',20)).pack()
            root6.mainloop()
            break
        else:
            msgbox.showinfo('提示','非法输入')
            break
def pound():
    rmb=RMB_entry.get()
    while True:
        try:
            rmb=float(rmb)
            rmb=round(rmb,2)
        except:
            msgbox.showinfo('提示','非法输入')
            break   
        if rmb>0:
            root7=Tk()
            root7.title("英镑")
            root7.geometry("300x150")
            root7.resizable(0,0)
            ammount=float(rmb)/8.10
            Label_dollar = Label(root7, text="可兑换%.2f英镑"%(ammount),font=('宋体',20))
            Label_dollar.pack()
            Button(root7, text="确定", command=root7.destroy,font=('宋体',20)).pack()
            root7.mainloop()
            break
        else:
            msgbox.showinfo('提示','非法输入')
            break
def yen():
    rmb=RMB_entry.get()
    while True:
        try:
            rmb=float(rmb)
            rmb=round(rmb,2)
        except:
            msgbox.showinfo('提示','非法输入')
            break   
        if rmb>0:
            root8=Tk()
            root8.title("日元")
            root8.geometry("300x150")
            root8.resizable(0,0)
            ammount=float(rmb)*20.17
            Label_dollar = Label(root8, text="可兑换%.2f日元"%(ammount),font=('宋体',20))
            Label_dollar.pack()
            Button(root8, text="确定", command=root8.destroy,font=('宋体',20)).pack() 
            root8.mainloop()
            break
        else:
            msgbox.showinfo('提示','非法输入')
            break

#登录判断
def reg():
    global a,r,i
    a=a_entry.get()
    b=p_entry.get()
    if len(a)==0:
        msgbox.showinfo('提示','请输入卡号')
    elif len(b)==0:
        msgbox.showinfo('提示','请输入密码')
    else:
        for x in ws['B']:
            if a==x.value:
                r=x.row
                Na=ws.cell(r,1).value
                if b==ws.cell(r,3).value:
                    msgbox.showinfo("提示","登录成功,欢迎%s"%(Na))
                    for x in root.place_slaves():
                        x.place_forget()
                    cent(root,402,216)
                    lab1.place(x=0,y=0)
                    root.after(0,update1,0)
                    break  
                else:
                    msgbox.showinfo("提示","密码错误")
                    break
        else:
            msgbox.showinfo("提示","卡号不存在")

#初始界面生成
def rot1():
    global root,lab3,a_entry,a_label,p_entry,p_label,btn,z,c,lab1,lab2,frames1,frames3,i
    root = Tk()
    i=1
    root.configure(bg='#f9f1f4')
    root.title('欢迎登录')
    root.resizable(0,0)
    cent(root,804,392)
    lab3=Label(root)
    c=Label(text="重庆大学模拟银行欢迎您!",font=('宋体',15),bg='#f9f1f4',fg='black')
    z=Label(text="提示:密码为6位数字!",font=('宋体',15),bg='#f9f1f4',fg='black')
    # 卡号
    a_label = Label(root,font=('宋体',25),text = "卡号：",bg='#f9f1f4',fg='black')
    a_entry = Entry(root,font=('宋体',25),bg='#f9f1f4',fg='black')
    # 密码
    p_entry = Entry(root,font=('宋体',25),bg='#f9f1f4',fg='black')
    p_entry["show"] = "*" # 密码显示为 *
    p_label = Label(root, text = "密码：",font=('宋体',25),bg='#f9f1f4',fg='black')
    # 登录按钮
    btn = Button(root, text = "登录", command = reg,font=('宋体',20))
    lab1=Label(root)
    lab2=Label(root)
    #开屏动画
    frames1 = get_gif(42,0,'login1.gif')
    frames3=get_gif(65,24,'login3(1).gif',master=root)
rot1()

frames2=get_gif(135,0,'login2(1).gif',master=root)
numIdx1 = 42 # gif的帧数
numIdx2=135
numIdx3=65

def update1(idx): # 定时器函数
    frame = frames1[idx]
    idx += 1 # 下一帧的序号
    lab1.configure(image=frame) # 显示当前帧的图片
    if idx==numIdx1-2:
        login()
        root.after(50, update1, idx%numIdx1) # 0.1秒(100毫秒)之后继续执行定时器函数(update)
    else:
        root.after(50, update1, idx%numIdx1)

def update2(idx): # 定时器函数
    frame = frames2[idx]
    idx += 1 # 下一帧的序号
    lab2.configure(image=frame) # 显示当前帧的图片
    if idx==numIdx2:
        lab2.destroy()
        mroot()
    else:
        root.after(50,update2,idx)
lab2.place(x=0,y=0)
root.after(0,update2,0)

def update3(idx): # 定时器函数
    frame = frames3[idx]
    idx += 1 # 下一帧的序号
    lab3.configure(image=frame) # 显示当前帧的图片
    if i==1:
        root.after(50,update3,idx%numIdx3)

#初始界面激活
def mroot():
    lab3.place(x=0,y=0)
    root.after(0,update3,0)
    root.resizable(0,0)
    c.place(x=30,y=10)
    z.place(x=30,y=360)
    a_label.place(x=223,y=120)
    a_entry.place(x=328,y=120)
    p_label.place(x=223,y=210)
    p_entry.place(x=328,y=210)
    btn.place(x=328,y=285)
    # 注册按钮
    Button(root, text='注册', command=register,font=('宋体',20)).place(x=548,y=285)
    #汇率
    lb_exrate=Label(root, text='汇率',font=('宋体',20))
    lb_exrate.place(x=730,y=360)
    lb_exrate.bind('<Button-1>',exrate)

def allnum(x):  #检验字符串是否全为数字
    for l in x:
        if "0"<=l<="9":
            None
        else:
            return False
    else:
        if x!="":
            return True
        else:
            return False
   
#注册界面
def register():
    global root1,name_entry,password_entry,password1_entry 
    root1=Tk()
    root1.resizable(0,0)
    root1.geometry('500x300')
    root1.title('重大模拟银行注册界面')
    y=Label(root1,text="提示:密码为6位数字!",font=('宋体',15))
    y.place(x=30,y=260)
    
    name=Label(root1,font=('宋体',15),text = "请输入用户名：")
    name.place(x=30,y=40)
    name_entry = Entry(root1,font=('宋体',15))
    name_entry.place(x=180,y=40)
    password=Label(root1,font=('宋体',15),text = "请输入密码：")
    password.place(x=30,y=100)
    password_entry = Entry(root1,font=('宋体',15))
    password_entry["show"] = "*" # 密码显示为 *
    password_entry.place(x=180,y=100)
    password1=Label(root1,font=('宋体',15),text = "请确认密码：")
    password1.place(x=30,y=160)
    password1_entry = Entry(root1,font=('宋体',15))
    password1_entry["show"] = "*" # 密码显示为 *
    password1_entry.place(x=180,y=160)
    Button(root1, text='确认注册', command=judge,font=('宋体',20)).place(x=200,y=200)
    root1.mainloop()

#将注册信息存入Excel
def append1():
    global accnt
    sig='6010900'
    account_n=str(random())[2:11]
    accnt=sig+account_n
    while True:
        for x in ws['B']:
            if x.value[7:]==account_n:
                account_n=random()

                break
        else:
            break
    x=2
    while True: 
        if ws.cell(x,1).value==None:
            ws.cell(x,1).value=C
            ws.cell(x,2).value=accnt
            ws.cell(x,3).value=A
            ws.cell(x,4).value=0
            ws.cell(x,5).value=0
            ws.cell(x,6).value=0
            break
        else:
            x+=1
    while True:
        if ws1.cell(x,1).value==None:
            ws1.cell(x,1).value=C
            ws1.cell(x,2).value=0
            break
        else:
            x+=1
    wb.save('Mytest.xlsx')
    wb.close()
    root1.destroy() 

#判断是否注册成功
def judge():
    global A,C
    A=password_entry.get()
    B=password1_entry.get()
    C=name_entry.get()
    if len(C)==0 :
        msgbox.showinfo("提示","输入错误，用户名不能为空")
    elif len(A)==0:
        msgbox.showinfo("提示","输入错误，密码不能为空")
    elif len(B)==0 :
        msgbox.showinfo("提示","输入错误，密码不能为空")
    elif len(A)==6 and len(C)!=0:        
        if  allnum(A):
            if A==B :
                append1()
                msgbox.showinfo("提示","注册成功！\n您的卡号为: %s\n (您的卡号已复制到剪贴板)"%(accnt))
                pc.copy(accnt)
                #break
            elif A!=B :
                msgbox.showinfo("提示","密码不一致，请重新设置")
                password_entry.delete(0,'end')
                password1_entry.delete(0,'end')

        else:
            msgbox.showinfo("提示","请勿使用字母，中文或特殊字符")
            password_entry.delete(0,'end')
            password1_entry.delete(0,'end')
    elif len(A)!=0 and len(A)!=6:
        msgbox.showinfo("提示","密码需要为6位数字")
        password_entry.delete(0,'end')
        password1_entry.delete(0,'end')
    root1.mainloop()

#修改密码：
def changepassword(event):
    global xg_entry,root4
    root4=Tk()
    root4.resizable(0,0)
    root4.geometry('500x300')
    root4.title('重大模拟银行修改密码界面')
    xg=Label(root4,font=('宋体',15),text = "请输入新密码：")
    xg.place(x=30,y=40)
    xg_entry = Entry(root4,font=('宋体',15))
    xg_entry.place(x=180,y=40)
    Button(root4, text='确认修改密码', command=reg2,font=('宋体',20)).place(x=200,y=150)
def reg2():
    newp=xg_entry.get()
    if allnum(newp) and len(newp)==6:
        ws.cell(r,3).value=newp
        msgbox.showinfo("提示","修改成功")
        root4.destroy()
        wb.save('Mytest.xlsx')
        wb.close()
        
    else:
        msgbox.showinfo('提示','请输入6位数字')
        
#主界面
def login():
    global root2
    root2=Tk()
    root2.resizable(0,0)
    cent(root2,550,300)
    root.destroy()
    root2.title('欢迎进入重大模拟银行')
    lab4=Label(root2)
    if a=='0':
        def but():
            Button(root2, text='统计', command=adminT,font=('宋体',30),fg='white',bg='black').place(x=50,y=40)
            Button(root2, text='冻结', command=adfrowindow,font=('宋体',30),fg='white',bg='black').place(x=50,y=160)
            Button(root2, text='解冻',command=defrowindow,font=('宋体',30),fg='white',bg='black').place(x=350,y=40)
            lb3=Label(root2, text='退出',font=('宋体',20), fg='red',bg='black')
            lb3.place(x=480,y=250)
            lb3.bind('<Button-1>',cancl)
        frames4=get_gif(65,0,'admin(1).gif',root2)
        numIdx4=65
        image1=get_image(root2,'admin.jpg',357,800)
    else:
        frames4=get_gif(46,0,'user(1).gif',root2)
        numIdx4=46
        image1=get_image(root2,'user.jpeg',357,800)
        def but():
            if ws.cell(r,7).value==1:
                Label(root2,text='您的账户已被冻结, 请向工作人员咨询情况',font=('宋体',15)).place(x=30,y=10)
                Button(root2, text='存款', command=Asidewindow,font=('宋体',30)).place(x=50,y=120)
                Button(root2, text='查询', command=Inquirewindow,font=('宋体',30)).place(x=350,y=120)
            else:
                Button(root2, text='取款', command=Paywindow,font=('宋体',30)).place(x=50,y=40)
                Button(root2, text='存款', command=Asidewindow,font=('宋体',30)).place(x=50,y=160)
                Button(root2, text='转账', command=Transwindow,font=('宋体',30)).place(x=350,y=40)
                Button(root2, text='查询', command=Inquirewindow,font=('宋体',30)).place(x=350,y=160)

            lb1=Label(root2, text='修改密码',font=('宋体',20),fg='#52a3f6',bg='#ede9e8')
            lb1.place(x=20,y=250)#修改密码按钮
            lb1.bind('<Button-1>',changepassword)
            lb2=Label(root2, text='注销',font=('宋体',20), fg='red',bg='#ede9e8')
            lb2.place(x=480,y=250)
            lb2.bind('<Button-1>',cancl)
    def update4(idx):
        frame = frames4[idx]
        idx += 1 # 下一帧的序号
        lab4.configure(image=frame) # 显示当前帧的图片
        if idx==numIdx4:
            lab4.configure(image=image1)
            lab4.place(x=-150,y=-25)
            but()
        else:
            root.after(50,update4,idx%numIdx4)
    lab4.place(x=-150,y=-20)
    root2.after(0,update4,0)
    root2.mainloop()

#注销
def cancl(event):
    root2.destroy()
    rot1()
    mroot()
    lab3.place(x=0,y=0)
    root.after(0,update3,0)

#取款界面
def Paywindow():
    global money_entryP, msg_labelP,Pay
    Pay=Tk()
    Pay.resizable(0,0)
    Pay.title('取款')
    cent(Pay,550,300)
    Label(Pay, text='取款金额',font=('宋体',25)).place(x=30, y=50)
    money_entryP=Entry(Pay,font=('宋体',25))
    money_entryP.place(x=170, y=50)
    Button(Pay, text='确认', font=('宋体',25),command=Paid).place(x=70,y=100)#此处要command=Paid来实现功能
    msg_labelP=Label(Pay, font=('宋体',20),text='')
    msg_labelP.place(x=50,y=180)
    Pay.mainloop()#构造取款界面
def Paid():
    remain=ws.cell(r,6).value  #提取用户信息
    amount=money_entryP.get()
    while True:
        try:
            amount=float(amount)
            amount=round(amount,2)
        except:
            msg_labelP['text']="非法输入，请正确输入数字"
            money_entryP.delete(0,'end')
            break      
        if amount>remain or amount<=0:
            msg_labelP['text']='金额不能为0或负数,不能超出余额'
            money_entryP.delete(0,'end')
            break
        elif 0<amount<=remain:
            msg_labelP['text']="取款成功, 余额%.2f"%(remain-amount)   
            money_entryP.delete(0,'end')
            ws.cell(r,6).value-=amount
            ws.cell(r,5).value+=amount
            ws.cell(2,5).value+=amount
            ws.cell(2,6).value-=amount
            x=2
            while True: 
                if ws1.cell(r,x).value==None:
                    ws1.cell(r,x).value=-amount
                    break
                else:
                    x+=1
            wb.save('Mytest.xlsx')
            wb.close()
            Button(Pay,text='返回', font=('宋体',25),command=Pay.destroy).place(x=200, y=100)
            break

#存款界面
def Asidewindow():
    global money_entryA, msg_labelA, Aside
    Aside=Tk()
    Aside.resizable(0,0)
    Aside.title('存款')
    cent(Aside,550,300)
    Label(Aside, text='存款金额',font=('宋体',25)).place(x=30,y=50)
    money_entryA=Entry(Aside,font=('宋体',25))
    money_entryA.place(x=170,y=50)
    Button(Aside, text='确认', font=('宋体',25),command=PutAside).place(x=70,y=100)#此处要command新的函数来实现功能
    msg_labelA=Label(Aside, font=('宋体',20),text='')
    msg_labelA.place(x=50,y=180)
    Aside.mainloop()#构造取款界面
def PutAside():
    remain=ws.cell(r,6).value    #找到用户所在行，提取参数
    amountA=money_entryA.get()
    la=len(amountA)
    while True:
        try:
            amountA=float(amountA)
            amountA=round(amountA,2)
        except:
            msg_labelA['text']="非法输入，请正确输入数字"
            money_entryA.delete(0,la)
            break   
        if amountA<=0:
            msg_labelA['text']='金额不能为0或负数' 
            money_entryA.delete(0,la)
            break
        else:       
            msg_labelA['text']="存款成功，余额%.2f"%(remain+amountA)  
            money_entryA.delete(0,la)
            ws.cell(r,6).value+=amountA
            ws.cell(r,4).value+=amountA
            ws.cell(2,4).value+=amountA
            ws.cell(2,6).value+=amountA
            x1=2
            while True: 
                if ws1.cell(r,x1).value==None:
                    ws1.cell(r,x1).value=amountA
                    break
                else:
                    x1+=1
            wb.save('Mytest.xlsx')
            wb.close()
            Button(Aside,text='返回', font=('宋体',25),command=Aside.destroy).place(x=200, y=100)
            break

#转账界面
def Transwindow():
    global money_entryT, Yaccount, msg_labelT,Transw
    Transw=Tk()
    Transw.resizable(0,0)
    Transw.title('转账')
    cent(Transw,550,300)
    Label(Transw, text='转给(卡号)', font=('宋体',20)).place(x=30,y=50)
    Yaccount=Entry(Transw,font=('宋体',25))
    Yaccount.place(x=170,y=50)
    Label(Transw, text='金额',font=('宋体',20)).place(x=30,y=120)
    money_entryT=Entry(Transw,font=('宋体',25))
    money_entryT.place(x=170,y=120)
    Button(Transw, text='确认', font=('宋体',25),command=Trans).place(x=170,y=170)#此处要command新的函数来实现功能
    msg_labelT=Label(Transw, font=('宋体',25),text='')
    msg_labelT.place(x=70,y=240)
    Transw.mainloop()#构造取款界面
def Trans():   
    amountT=money_entryT.get()
    You=Yaccount.get()
    remain=ws.cell(r,6).value    #找到用户所在行，提取参数
    for z in ws['B']:
                if You==z.value and You !=a and You!='0':
                    y=z.row
                    while True:
                        try:
                            amountT=float(amountT)
                            amountT=round(amountT,2)
                        except:
                            msg_labelT['text']="非法输入，请正确输入数字"
                            money_entryT.delete(0,'end')
                            break           
                        if 0<amountT<=remain:
                            msg_labelT['text']="转账成功，余额%.2f"%(remain-amountT)   
                            money_entryT.delete(0,'end')
                            Yaccount.delete(0,len(You))
                            ws.cell(r,6).value-=amountT
                            ws.cell(r,5).value+=amountT
                            ws.cell(y,6).value+=amountT
                            ws.cell(y,4).value+=amountT
                            x2=2
                            while True: 
                                if ws1.cell(r,x2).value==None:
                                    ws1.cell(r,x2).value=-amountT
                                    break
                                else:
                                    x2+=1
                            x3=2
                            while True:
                                if ws1.cell(y,x3).value==None:
                                    ws1.cell(y,x3).value=amountT
                                    break
                                else:
                                    x3+=1
                            wb.save('Mytest.xlsx')
                            wb.close()
                            Button(Transw,text='返回',font=('宋体',25),command=Transw.destroy).place(x=300,y=170)
                            break
                        else:
                            msg_labelT['text']='请正确输入金额'
                            money_entryT.delete(0,'end')
                            break
                    break
    else:
        msg_labelT['text']='请不要给自己或不存在的人转账'
        Yaccount.delete(0,'end')

#查询界面
def Inquirewindow():
    inq=Tk()
    inq.resizable(0,0)
    inq.title('查询')
    an=ws[r]
    tree=Treeview(inq,columns=('收入','支出','存款','状态'))
    tree.heading('#0',text='用户名')
    tree.heading('#1',text='收入')
    tree.heading('#2',text='支出')
    tree.heading('#3',text='存款')
    tree.heading('#4',text='状态')
    if an[6].value==1:
        st='冻结'
    else:
        st='正常'
    tree.insert('',text=an[1].value,index=tkinter.END,values=(an[3].value,an[4].value,an[5].value,st))
    tree.grid()
    Button(inq,text='统计',command=total).grid(row=5, sticky=E)
    inq.mainloop()

#统计界面
def total():
    data=[]
    for i in ws1[r]:
        data.append(i.value)
    data1=data[2::]
    data2=[]
    for i in data1:
        if i!=None:
            data2.append(i)
    money=data[1]
    data3=[money]
    for i in data2:
        money+=i
        data3.append(money)
    wb.close()

    #以下用于绘制图表
    import matplotlib.pyplot as plt
    import numpy as np
    plt.rcParams['font.sans-serif']=['Simhei']
    x=[]
    z=0
    for i in data3:
        x.append(z)
        z+=1
    plt.plot(x, data3, "g", marker='D', markersize=5, label="余额")
    plt.xlabel("交易编号")
    plt.ylabel("余额")
    plt.title("账户余额变化记录")
    plt.legend(loc="lower right")
    #调用 text()在图像上绘制注释文本
    #x1、y1表示文本所处坐标位置，ha参数控制水平对齐方式, va控制垂直对齐方式，str(y1)表示要绘制的文本
    for x1, y1 in zip(x, data3):
        plt.text(x1, y1, str(y1), ha='center', va='bottom', fontsize=10)
    x0=np.arange(x[0], x[-1]+1, 1) #设置x刻度
    plt.xticks(x0)
    plt.savefig("1.jpg")
    plt.show()
def adminT():
    moneyin=ws.cell(2,4).value
    moneyout=ws.cell(2,5).value
    moneyrm=ws.cell(2,6).value
    wb.close()
    #读取收入、支出、余额（总）
    import matplotlib.pyplot as plt
    plt.rcParams['font.sans-serif']=['Simhei']
    fig,ax=plt.subplots()
    ax.set_ylabel('资金（￥）')
    ax.set_title('银行账户总收入,总支出和总余额')
    p1=ax.bar('总收入',moneyin,0.35)
    p2=ax.bar('总支出',moneyout,0.35)
    p3=ax.bar('总额度',moneyrm,0.35)
    ax.legend((p1[0],p2[0],p3[0]),('总支出','总收入','总额度'))
    plt.show()

#冻结
def adfrowindow():
    global lsf,lsf1,top
    top = Tk()
    top.resizable(0,0)
    top.title('冻结')
    top.geometry('500x450')
    Label(top,text='请选择要冻结的账户').grid(row=0,sticky=W)
    lsf=[]
    lsf1=[]
    for x in ws['B']:
        r=x.row
        if x.value!='卡号' and x.value!='0':
            exec('chkVar%s=IntVar(master=top)'%r)
            exec('C%s=Checkbutton(top, text = "%s   %s",variable=chkVar%s,\
                onvalue=1,offvalue=0)'%(r,ws.cell(r,1).value,x.value,r))
            exec('C%s.grid(sticky=W)'%(r))
            exec('lsf.append(chkVar%s)'%r)
    Button(top,text='确认', command=froze).place(x=400, y=400)
    top.mainloop()
def froze():
    for x in lsf:
        lsf1.append(x.get())
    for y in range(0,len(lsf1)):
        if lsf1[y]==1:
            ws.cell(y+3,7).value=1
            wb.save('Mytest.xlsx')
            wb.close()
    msgbox.showinfo('提示','冻结成功')
    top.destroy()

#解冻
def defrowindow():
    global topdeFr,dic,dic1
    topdeFr=Tk()
    topdeFr.resizable(0,0)
    topdeFr.title('解冻')
    topdeFr.geometry('500x450')
    Label(topdeFr,text='请选择要解冻的账户').grid(row=0,sticky=W)
    dic={}
    dic1={}
    for x in ws['B']:
        r=x.row
        if ws.cell(r,7).value==1:
            exec('chkVar%s=IntVar(master=topdeFr)'%r)
            exec('C%s=Checkbutton(topdeFr, text = "%s   %s",variable=chkVar%s,\
                onvalue=1,offvalue=0)'%(r,ws.cell(r,1).value,x.value,r))
            exec('C%s.grid(sticky=W)'%(r))
            exec('dic[%s]=chkVar%s'%(r,r))
    Button(topdeFr,text='确认', command=deFroze).place(x=400, y=400)
    topdeFr.mainloop()
def deFroze():
    for x in dic.keys():
        dic1[x]=(dic[x].get())
    for y in dic1.keys():
        if dic1[y]==1:
            ws.cell(y,7).value=None
            wb.save('Mytest.xlsx')
            wb.close()
    msgbox.showinfo('提示','解冻成功')
    topdeFr.destroy()

root.mainloop()